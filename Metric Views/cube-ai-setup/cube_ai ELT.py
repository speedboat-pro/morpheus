# Databricks notebook source
# MAGIC %md
# MAGIC # Notebook Setup

# COMMAND ----------

# MAGIC %pip install openpyxl
# MAGIC dbutils.library.restartPython()
# MAGIC

# COMMAND ----------

from openpyxl import load_workbook
import pyspark.sql.functions as F
spark.conf.set("spark.sql.ansi.enabled", False)
import pyspark.pandas as ps

# COMMAND ----------

cat = dbutils.widgets.get('cat')
spark.catalog.setCurrentCatalog(f'{cat}')
db = dbutils.widgets.get('db')
spark.catalog.setCurrentDatabase(f'{db}')

# COMMAND ----------

# Loading an Excel workbook
wb = load_workbook(f'/Volumes/{cat}/{db}/diad/USSales/bi_dimensions.xlsx', data_only=True, read_only=True)
for w in wb: print (w.title)

# COMMAND ----------

# MAGIC %md
# MAGIC #M (Power Query) & DAX

# COMMAND ----------

# MAGIC %md
# MAGIC ## Fact Table (M)

# COMMAND ----------

# MAGIC %md
# MAGIC ### Sales
# MAGIC
# MAGIC ```
# MAGIC
# MAGIC # let
# MAGIC #    Source = Csv.Document(File.Contents("C:\DIAD\Data\USSales\Sales.csv"),[Delimiter=",", Columns=5, Encoding=1252, QuoteStyle=QuoteStyle.None]),
# MAGIC     #"Promoted Headers" = Table.PromoteHeaders(Source, [PromoteAllScalars=true]),
# MAGIC     #"Changed Type" = Table.TransformColumnTypes(#"Promoted Headers",{{"ProductID", Int64.Type}, {"Date", type date}, {"Zip", type text}, {"Units", Int64.Type}, {"Revenue", type number}}),
# MAGIC     #"Appended Query" = Table.Combine({#"Changed Type", International}),
# MAGIC     #"Added Conditional Column" = Table.AddColumn(#"Appended Query", "Custom", each if [Country] = null then "USA" else [Country]),
# MAGIC     #"Removed Columns" = Table.RemoveColumns(#"Added Conditional Column",{"Country"}),
# MAGIC     #"Renamed Columns" = Table.RenameColumns(#"Removed Columns",{{"Custom", "Country"}})
# MAGIC #in
# MAGIC     #"Renamed Columns"
# MAGIC
# MAGIC ```

# COMMAND ----------

# MAGIC %md
# MAGIC ### _International_
# MAGIC ```
# MAGIC let
# MAGIC     Source = Folder.Files("C:\DIAD\Data\InternationalSales"),
# MAGIC     #"Invoke Custom Function1" = Table.AddColumn(Source, "Transform File from International", each #"Transform File from International"([Content])),
# MAGIC     #"Renamed Columns1" = Table.RenameColumns(#"Invoke Custom Function1", {"Name", "Source.Name"}),
# MAGIC     #"Removed Other Columns1" = Table.SelectColumns(#"Renamed Columns1", {"Source.Name", "Transform File from International"}),
# MAGIC     #"Expanded Table Column1" = Table.ExpandTableColumn(#"Removed Other Columns1", "Transform File from International", Table.ColumnNames(#"Transform File from International"(#"Sample File"))),
# MAGIC     #"Changed Type" = Table.TransformColumnTypes(#"Expanded Table Column1",{{"Source.Name", type text}, {"ProductID", Int64.Type}, {"Date", type date}, {"Zip", type text}, {"Units", Int64.Type}, {"Revenue", type number}, {"Country", type text}}),
# MAGIC     #"Removed Columns" = Table.RemoveColumns(#"Changed Type",{"Source.Name"})
# MAGIC in
# MAGIC     #"Removed Columns"
# MAGIC ```

# COMMAND ----------

# MAGIC %md
# MAGIC ### Transform File from International\Sample Query\ _Sample File Parameter1 (Sample File)_
# MAGIC ```
# MAGIC #"Sample File" meta [IsParameterQuery=true, BinaryIdentifier=#"Sample File", Type="Binary", IsParameterQueryRequired=true]
# MAGIC ```

# COMMAND ----------

# MAGIC %md
# MAGIC ### Transform File from International\ _Transform File from International_
# MAGIC ```
# MAGIC let
# MAGIC     Source = (#"Sample File Parameter1" as binary) => let
# MAGIC     Source = Csv.Document(#"Sample File Parameter1",[Delimiter=",", Columns=6, Encoding=1252, QuoteStyle=QuoteStyle.None]),
# MAGIC     #"Promoted Headers" = Table.PromoteHeaders(Source, [PromoteAllScalars=true])
# MAGIC in
# MAGIC     #"Promoted Headers"
# MAGIC in
# MAGIC     Source
# MAGIC ```

# COMMAND ----------

# MAGIC %md
# MAGIC ### Transform File from International\ _Transform Sample File from International_
# MAGIC ```
# MAGIC let
# MAGIC     Source = Csv.Document(#"Sample File Parameter1",[Delimiter=",", Columns=6, Encoding=1252, QuoteStyle=QuoteStyle.None]),
# MAGIC     #"Promoted Headers" = Table.PromoteHeaders(Source, [PromoteAllScalars=true])
# MAGIC in
# MAGIC     #"Promoted Headers"
# MAGIC ```

# COMMAND ----------

# MAGIC %md
# MAGIC ### Transform File from International\Sample Query\ _Sample File_
# MAGIC ```
# MAGIC let
# MAGIC     Source = Folder.Files("C:\DIAD\Data\InternationalSales"),
# MAGIC     Navigation1 = Source{0}[Content]
# MAGIC in
# MAGIC     Navigation1
# MAGIC ```

# COMMAND ----------

# MAGIC %md
# MAGIC ## Dimension Tables

# COMMAND ----------

# MAGIC %md
# MAGIC ### Date (DAX)

# COMMAND ----------

# MAGIC %md
# MAGIC ```
# MAGIC Date = CALENDAR (DATE(2013,1,1), DATE(2021,12,31))
# MAGIC ```

# COMMAND ----------

# MAGIC %md
# MAGIC ### Geography

# COMMAND ----------

# MAGIC %md
# MAGIC #### M

# COMMAND ----------

# MAGIC %md
# MAGIC ```
# MAGIC let
# MAGIC     Source = Excel.Workbook(File.Contents("C:\DIAD\Data\USSales\bi_dimensions.xlsx"), null, true),
# MAGIC     Geo_Sheet = Source{[Item="geo",Kind="Sheet"]}[Data],
# MAGIC     #"Changed Type" = Table.TransformColumnTypes(Geo_Sheet,{{"Column1", type text}, {"Column2", type any}, {"Column3", type text}, {"Column4", type text}, {"Column5", type text}, {"Column6", type text}}),
# MAGIC     #"Removed Top Rows" = Table.Skip(#"Changed Type",3),
# MAGIC     #"Promoted Headers" = Table.PromoteHeaders(#"Removed Top Rows", [PromoteAllScalars=true]),
# MAGIC     #"Changed Type1" = Table.TransformColumnTypes(#"Promoted Headers",{{"Zip", type text}, {"City", type text}, {"State", type text}, {"Region", type text}, {"District", type text}, {"Country", type text}})
# MAGIC in
# MAGIC     #"Changed Type1"
# MAGIC ```

# COMMAND ----------

# MAGIC %md
# MAGIC #### DAX

# COMMAND ----------

# MAGIC %md
# MAGIC ```
# MAGIC ZipCountry = Geography[Zip]&"|"&Geography[Country]
# MAGIC ```

# COMMAND ----------

# MAGIC %md
# MAGIC ### Manufacturer (M)
# MAGIC
# MAGIC ```
# MAGIC let
# MAGIC     Source = Excel.Workbook(File.Contents("C:\DIAD\Data\USSales\bi_dimensions.xlsx"), null, true),
# MAGIC     manufacturer_Sheet = Source{[Item="manufacturer",Kind="Sheet"]}[Data],
# MAGIC     #"Promoted Headers" = Table.PromoteHeaders(manufacturer_Sheet, [PromoteAllScalars=true]),
# MAGIC     #"Changed Type" = Table.TransformColumnTypes(#"Promoted Headers",{{"Column1", type text}, {"Column2", type any}, {"Column3", type any}, {"Column4", type any}, {"Column5", type any}, {"Column6", type any}, {"Column7", type any}, {"Column8", type any}, {"Column9", type any}, {"Column10", type any}, {"Column11", type any}, {"Column12", type any}, {"Column13", type any}, {"Column14", type any}, {"Column15", type any}}),
# MAGIC     #"Removed Bottom Rows" = Table.RemoveLastN(#"Changed Type",3),
# MAGIC     #"Transposed Table" = Table.Transpose(#"Removed Bottom Rows"),
# MAGIC     #"Promoted Headers1" = Table.PromoteHeaders(#"Transposed Table", [PromoteAllScalars=true]),
# MAGIC     #"Changed Type1" = Table.TransformColumnTypes(#"Promoted Headers1",{{"ManufacturerID", Int64.Type}, {"Manufacturer", type text}, {"Logo", type text}})
# MAGIC in
# MAGIC     #"Changed Type1"
# MAGIC ```

# COMMAND ----------

# MAGIC %md
# MAGIC ### Product (M)
# MAGIC
# MAGIC ```
# MAGIC let
# MAGIC     Source = Excel.Workbook(File.Contents("C:\DIAD\Data\USSales\bi_dimensions.xlsx"), null, true),
# MAGIC     Product_Table = Source{[Item="Product_Table",Kind="Table"]}[Data],
# MAGIC     #"Changed Type" = Table.TransformColumnTypes(Product_Table,{{"ProductID", Int64.Type}, {"Product", type text}, {"Category", type text}, {"ManufacturerID", Int64.Type}, {"Price", type text}}),
# MAGIC     #"Inserted Text Before Delimiter" = Table.AddColumn(#"Changed Type", "Text Before Delimiter", each Text.BeforeDelimiter([Product], "|"), type text),
# MAGIC     #"Inserted Text After Delimiter" = Table.AddColumn(#"Inserted Text Before Delimiter", "Text After Delimiter", each Text.AfterDelimiter([Product], "|"), type text),
# MAGIC     #"Removed Columns" = Table.RemoveColumns(#"Inserted Text After Delimiter",{"Product"}),
# MAGIC     #"Renamed Columns" = Table.RenameColumns(#"Removed Columns",{{"Text Before Delimiter", "Product"}, {"Text After Delimiter", "Segment"}}),
# MAGIC     #"Filled Down" = Table.FillDown(#"Renamed Columns",{"Category"})
# MAGIC in
# MAGIC     #"Filled Down"
# MAGIC ```

# COMMAND ----------

# MAGIC %md
# MAGIC #  PySpark / Spark SQL 

# COMMAND ----------

# MAGIC %md
# MAGIC ## Fact Table

# COMMAND ----------

# MAGIC %md
# MAGIC ### Sales

# COMMAND ----------

spark.sql(
  f'''
    CREATE OR REPLACE TABLE {cat}.{db}.factsales
    SELECT
    ProductID	
    ,Date	
    ,CAST(Zip AS STRING) Zip
    ,Units	
    ,Revenue	
    ,'USA' as Country
    , concat(Zip, '@USA') ZipCountry
    FROM
      read_files('/Volumes/{cat}/{db}/diad/USSales/Sales.csv')
    UNION ALL
    SELECT
      ProductID	
      ,Date	
      ,CAST(Zip AS STRING)	
      ,Units	
      ,Revenue	
      ,Country
      , concat(Zip, '@', Country) ZipCountry

    FROM
      read_files('/Volumes/{cat}/{db}/diad/InternationalSales')'''
)

# COMMAND ----------

# MAGIC %md
# MAGIC ## Dimension Tables

# COMMAND ----------

# MAGIC %md
# MAGIC ### Date

# COMMAND ----------

# spark.sql(f'''
#     CREATE OR REPLACE TABLE {cat}.{db}.dimdate
#       SELECT
#         explode(
#           sequence(
#             to_date('2013-01-01'),
#             to_date('2021-12-31'),
#             INTERVAL 1 DAY
#           )
#         ) AS date
#   ''')

# COMMAND ----------

# MAGIC %md
# MAGIC ### Geography

# COMMAND ----------

rows = [r for r in wb['geo'].values]
columns = [[x for x in n] for n in rows[1:]]
(
  spark.createDataFrame(
    columns[3:]).toDF(*columns[2])
    .withColumn('ZipCountry', F.concat(F.col('Zip'), F.lit('@'),F.col('Country')))
    .write.mode("overwrite").format("delta").option("delta.columnMapping.mode", "name").option("mergeSchema", "true").saveAsTable("dimgeography")
  )

# COMMAND ----------

# MAGIC %md
# MAGIC ### Manufacturer

# COMMAND ----------

rows = [r for r in wb['manufacturer'].values]
columns = [[x for x in n] for n in rows[1:]]
transposed = list(map(list, zip(*columns[0:3])))
spark.createDataFrame(transposed[1:]).toDF(*transposed[0]).write.mode("overwrite").format("delta").option("delta.columnMapping.mode", "name").saveAsTable("dimmanufacturer")

# COMMAND ----------

# MAGIC %md
# MAGIC ### Product

# COMMAND ----------

rows = [r for r in wb['product'].values]
columns = [[x for x in n] for n in rows[1:]]
prod_df = spark.createDataFrame(columns[1:]).toDF(*columns[0]).write.mode("overwrite").option("delta.columnMapping.mode", "name").option("mergeSchema", "true").saveAsTable("dimproduct")

# COMMAND ----------

prod_df = ps.DataFrame(
  spark.sql(
    f'''
    SELECT
        Category
      , ManufacturerID  
      , Price
      , split_part(product, '|', 1) AS Product
      , ProductID
      , split_part(product, '|', 2) AS Segment
    FROM 
      {cat}.{db}.dimproduct
      ''')).ffill().to_spark().write.mode("overwrite").option("delta.columnMapping.mode", "name").option("mergeSchema", "true").saveAsTable("dimproduct")


